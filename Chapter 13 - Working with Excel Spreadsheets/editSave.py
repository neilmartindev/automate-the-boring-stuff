import openpyxl

wb = openpyxl.load_workbook('testsheet.xlsx')

sheet = wb.active

sheet.title = 'Spam spam spam'

wb.create_sheet(index=1, title='First Sheet')

wb.create_sheet(index=2, title='Middle Sheet')

wb.save('testsheetcopy.xlsx')

wb.sheetnames
